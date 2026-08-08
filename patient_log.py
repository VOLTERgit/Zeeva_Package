"""
Zeeva Clinic - Patient Excel Log
--------------------------------
Every time staff generates a package PDF, one row gets added automatically
to Zeeva_Patients_Log.xlsx (created next to the app the first time it's
needed). Staff never has to touch this - it just builds up over time as a
running record of every package generated.
"""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

HEADERS = [
    "Date",
    "Patient Name",
    "Age",
    "Phone Number",
    "Country",
    "Technique",
    "Services Included",
    "Total (Rs)",
    "GST Applied",
    "PDF Filename",
    "Logged At",
]


def _new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Patients"
    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 4)
    return wb, ws


def log_patient_to_excel(log_path, patient, items_included, technique, total,
                          gst_applied, pdf_filename):
    """
    Append one row for this patient/package to the Excel log at log_path.
    Creates the file (with headers) if it doesn't exist yet.
    """
    if os.path.exists(log_path):
        wb = load_workbook(log_path)
        ws = wb["Patients"] if "Patients" in wb.sheetnames else wb.active
    else:
        wb, ws = _new_workbook()

    services_str = ", ".join(item["name"] for item in items_included)

    ws.append([
        patient.get("date", ""),
        patient.get("name", ""),
        patient.get("age", ""),
        patient.get("phone", ""),
        patient.get("country", ""),
        technique,
        services_str,
        round(total, 2),
        "Yes" if gst_applied else "No",
        pdf_filename,
        datetime.now().strftime("%d %b %Y %H:%M"),
    ])

    wb.save(log_path)
